"""
数据比对核心模块
实现订单数据与产品规格数据的比对逻辑
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
import os
import uuid
import logging

# 设置日志记录器
logger = logging.getLogger(__name__)

class OrderSpecComparator:
    def __init__(self, output_dir='outputs'):
        # 确保使用绝对路径
        if not os.path.isabs(output_dir):
            # 获取项目根目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            self.output_dir = os.path.join(project_root, output_dir)
        else:
            self.output_dir = output_dir
        self.ensure_output_dir()
        
        # 定义错误类型
        self.ERROR_TYPES = {
            'PRODUCT_NOT_FOUND': '产品ID不存在',
            'SIZE_MISMATCH': '尺寸不符',
            'COLOR_MISMATCH': '颜色不符',
            'PRICE_MISMATCH': '单价不符',
            'TOTAL_CALC_ERROR': '总价计算错误'
        }
        
        # 定义高亮颜色
        self.ERROR_FILL = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        self.ERROR_FONT = Font(color='CC0000', bold=True)
        
    def ensure_output_dir(self):
        """确保输出目录存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def _infer_columns_by_content(self, df):
        """基于数据内容推断列名"""
        column_mapping = {}
        
        for col in df.columns:
            # 获取该列的非空值样本
            sample_values = df[col].dropna().astype(str).head(10).tolist()
            if not sample_values:
                continue
                
            # 分析数据模式
            sample_text = ' '.join(sample_values).lower()
            
            # 检查是否是产品ID列（通常包含字母数字组合）
            if any(val for val in sample_values if len(str(val)) > 3 and any(c.isalpha() for c in str(val)) and any(c.isdigit() for c in str(val))):
                if 'item_id' not in column_mapping.values():
                    column_mapping[col] = 'item_id'
                    continue
            
            # 检查是否是价格列（数字且通常有小数点）
            numeric_count = sum(1 for val in sample_values if self._is_numeric(val))
            if numeric_count > len(sample_values) * 0.7:  # 70%以上是数字
                decimal_count = sum(1 for val in sample_values if '.' in str(val))
                if decimal_count > 0 and 'unit_price' not in column_mapping.values():
                    column_mapping[col] = 'unit_price'
                    continue
                elif 'quantity' not in column_mapping.values():
                    column_mapping[col] = 'quantity'
                    continue
            
            # 检查是否是产品名称列（通常是中文或较长的文本）
            text_count = sum(1 for val in sample_values if len(str(val)) > 2 and not self._is_numeric(val))
            if text_count > len(sample_values) * 0.5 and 'product_name' not in column_mapping.values():
                column_mapping[col] = 'product_name'
                continue
        
        return column_mapping
    
    def _is_numeric(self, value):
        """检查值是否为数字"""
        try:
            float(str(value).replace(',', ''))
            return True
        except (ValueError, TypeError):
            return False
    
    def _get_price_value(self, row):
        """获取行中的价格值，尝试多个可能的列名"""
        price_columns = ['unit_price', 'price', '单价', '价格', 'cost']
        
        for col in price_columns:
            if col in row and not pd.isna(row[col]):
                try:
                    price = float(row[col])
                    if price > 0:
                        return price
                except (ValueError, TypeError):
                    continue
        
        return 0.0
    
    def _get_quantity_value(self, row):
        """获取行中的数量值"""
        quantity_columns = ['quantity', 'qty', '数量', '件数']
        
        for col in quantity_columns:
            if col in row and not pd.isna(row[col]):
                try:
                    qty = float(row[col])
                    if qty > 0:
                        return qty
                except (ValueError, TypeError):
                    continue
        
        return 0.0
    
    def _get_total_value(self, row):
        """获取行中的总价值"""
        total_columns = ['total_price', 'total', '总价', '合计', '小计']
        
        for col in total_columns:
            if col in row and not pd.isna(row[col]):
                try:
                    total = float(row[col])
                    if total > 0:
                        return total
                except (ValueError, TypeError):
                    continue
        
        return 0.0
    
    def _has_quantity_and_total(self, row):
        """检查行中是否有数量和总价数据"""
        return self._get_quantity_value(row) > 0 and self._get_total_value(row) > 0
            
    def load_order_data(self, order_file_path):
        """
        加载订单数据，支持多个工作表
        
        Args:
            order_file_path: 订单Excel文件路径
            
        Returns:
            pandas.DataFrame: 合并后的订单数据，如果失败返回None
        """
        try:
            # 读取Excel文件的所有工作表
            excel_file = pd.ExcelFile(order_file_path)
            sheet_names = excel_file.sheet_names
            
            logger.info(f"发现 {len(sheet_names)} 个工作表: {sheet_names}")
            
            all_dataframes = []
            
            # 遍历所有工作表
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(order_file_path, sheet_name=sheet_name)
                    
                    # 跳过空工作表
                    if df.empty:
                        logger.info(f"跳过空工作表: {sheet_name}")
                        continue
                    
                    # 添加工作表标识列
                    df['工作表'] = sheet_name
                    df['表格序号'] = len(all_dataframes) + 1
                    
                    all_dataframes.append(df)
                    logger.info(f"成功加载工作表 '{sheet_name}': {len(df)} 行数据")
                    
                except Exception as e:
                    logger.warning(f"加载工作表 '{sheet_name}' 失败: {str(e)}")
                    continue
            
            if not all_dataframes:
                logger.error("没有成功加载任何工作表")
                return None
            
            # 合并所有工作表的数据
            df = pd.concat(all_dataframes, ignore_index=True)
            logger.info(f"合并后总数据行数: {len(df)}")
            
            # 如果只有一个工作表，移除工作表标识列以保持兼容性
            if len(all_dataframes) == 1:
                df = df.drop(['工作表', '表格序号'], axis=1)
            
            # 数据清洗
            df = df.dropna(how='all')  # 删除完全空白的行
            
            # 标准化列名（去除空格，转换为小写）
            # 确保列名是字符串类型
            df.columns = [str(col).strip().lower() for col in df.columns]
            
            # 检查是否是从PDF转换的Excel文件（列名可能是数字）
            if all(col.isdigit() for col in df.columns if col.strip()):
                logger.info("检测到从PDF转换的Excel文件，尝试推断列名")
                
                # 尝试从前几行推断列名
                header_rows = df.head(5)  # 取前5行尝试推断列名
                
                # 查找可能的标题行
                potential_headers = {}
                for i, row in header_rows.iterrows():
                    # 检查这一行是否包含关键字
                    row_values = [str(v).lower() for v in row.values if pd.notna(v)]
                    keywords = ['item', 'product', 'price', 'size', 'color', 'quantity', 'amount']
                    if any(kw in ' '.join(row_values) for kw in keywords):
                        # 记录这一行可能包含的列名
                        for col_idx, value in enumerate(row):
                            if pd.notna(value) and str(value).strip():
                                col_name = str(col_idx)
                                if col_name not in potential_headers:
                                    potential_headers[col_name] = []
                                potential_headers[col_name].append(str(value).lower())
                
                # 创建映射字典
                column_mapping = {}
                
                # 尝试映射列名
                for col, values in potential_headers.items():
                    # 合并这一列的所有可能值
                    combined = ' '.join(values).lower()
                    
                    # 更强大的列名映射逻辑
                    if any(kw in combined for kw in ['item', 'number', 'id', '编号', '货号', '产品编号', '商品编号']):
                        column_mapping[col] = 'item_id'
                    elif any(kw in combined for kw in ['size', 'dimension', '尺寸', '规格', '型号']):
                        column_mapping[col] = 'size'
                    elif any(kw in combined for kw in ['color', 'colour', '颜色', '色彩']):
                        column_mapping[col] = 'color'
                    elif any(kw in combined for kw in ['price', 'unit', 'cost', '单价', '价格', '金额']):
                        column_mapping[col] = 'unit_price'
                    elif any(kw in combined for kw in ['quantity', 'qty', 'amount', '数量', '件数']):
                        column_mapping[col] = 'quantity'
                    elif any(kw in combined for kw in ['total', 'sum', '总价', '合计', '小计']):
                        column_mapping[col] = 'total_price'
                    elif any(kw in combined for kw in ['product', 'description', 'name', '产品', '商品', '名称', '描述']):
                        column_mapping[col] = 'product_name'
                
                # 如果找到了映射，重命名列并跳过标题行
                if column_mapping:
                    # 找到最后一个标题行的索引，但要确保不删除所有数据
                    last_header_row = max(header_rows.index)
                    
                    # 只有当标题行不是最后一行时才删除
                    if last_header_row < len(df) - 1:
                        df = df.iloc[last_header_row+1:].reset_index(drop=True)
                    else:
                        # 如果标题行是最后几行，只删除第一行
                        df = df.iloc[1:].reset_index(drop=True)
                    
                    # 重命名列
                    df = df.rename(columns=column_mapping)
                    
                    logger.info(f"列映射结果: {column_mapping}")
                    logger.info(f"删除标题行后剩余数据行数: {len(df)}")
                else:
                    logger.warning("无法推断列名，尝试基于数据内容推断")
                    # 尝试基于数据内容推断列名
                    column_mapping = self._infer_columns_by_content(df)
                    if column_mapping:
                        df = df.rename(columns=column_mapping)
                        logger.info(f"基于内容推断的列映射结果: {column_mapping}")
                    else:
                        logger.warning("完全无法推断列名，使用默认列名")
            
            # 检查必需的列是否存在（进一步放宽要求）
            required_columns = ['item_id']  # 只要求item_id为必需列
            optional_columns = ['unit_price', 'product_name', 'quantity', 'size', 'color']
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                logger.error(f"订单文件缺少必需的列: {missing_columns}")
                logger.info(f"当前列名: {list(df.columns)}")
                logger.info("尝试使用第一列作为item_id")
                
                # 如果只缺少item_id，尝试使用第一列
                if missing_columns == ['item_id'] and len(df.columns) > 0:
                    first_col = df.columns[0]
                    df = df.rename(columns={first_col: 'item_id'})
                    logger.info(f"将第一列 '{first_col}' 重命名为 'item_id'")
                else:
                    return None
                
            # 数据类型转换和清洗
            df['item_id'] = df['item_id'].fillna('').astype(str).str.strip()
            
            # 为缺少的列添加默认值
            for col in optional_columns:
                if col not in df.columns:
                    if col in ['unit_price', 'quantity', 'total_price']:
                        df[col] = 0.0  # 数字列默认为0
                    else:
                        df[col] = ''   # 文本列默认为空字符串
            
            # 数据类型转换
            if 'size' in df.columns:
                df['size'] = df['size'].fillna('').astype(str).str.strip()
                
            if 'color' in df.columns:
                df['color'] = df['color'].fillna('').astype(str).str.strip()
                
            if 'product_name' in df.columns:
                df['product_name'] = df['product_name'].fillna('').astype(str).str.strip()
                
            if 'unit_price' in df.columns:
                df['unit_price'] = pd.to_numeric(df['unit_price'], errors='coerce').fillna(0.0)
            
            if 'quantity' in df.columns:
                df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0.0)
            
            if 'total_price' in df.columns:
                df['total_price'] = pd.to_numeric(df['total_price'], errors='coerce').fillna(0.0)
            
            return df
            
        except Exception as e:
            logger.error(f"加载订单数据失败: {str(e)}")
            return None
            
    def load_spec_data(self, spec_file_path):
        """
        加载产品规格数据
        
        Args:
            spec_file_path: 产品规格Excel文件路径
            
        Returns:
            pandas.DataFrame: 产品规格数据，如果失败返回None
        """
        try:
            df = pd.read_excel(spec_file_path)
            
            # 数据清洗
            df = df.dropna(how='all')  # 删除完全空白的行
            
            # 标准化列名
            # 确保列名是字符串类型
            df.columns = [str(col).strip().lower() for col in df.columns]
            
            # 尝试映射常见的列名变体
            column_mapping = {}
            for col in df.columns:
                col_lower = col.lower()
                if any(kw in col_lower for kw in ['item', 'id', 'number', 'code']):
                    column_mapping[col] = 'item_id'
                elif any(kw in col_lower for kw in ['size', 'dimension']):
                    column_mapping[col] = 'size'
                elif any(kw in col_lower for kw in ['color', 'colour']):
                    column_mapping[col] = 'color'
                elif any(kw in col_lower for kw in ['price', 'unit', 'cost', 'standard']):
                    column_mapping[col] = 'standard_unit_price'
                elif any(kw in col_lower for kw in ['product', 'name', 'description']):
                    column_mapping[col] = 'product_name'
            
            # 应用映射
            if column_mapping:
                df = df.rename(columns=column_mapping)
                logger.info(f"规格表列映射结果: {column_mapping}")
            
            # 检查必需的列是否存在
            required_columns = ['item_id', 'standard_unit_price']  # 减少必需列，增加灵活性
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                logger.error(f"规格表文件缺少必需的列: {missing_columns}")
                return None
                
            # 数据类型转换和清洗
            df['item_id'] = df['item_id'].fillna('').astype(str).str.strip()
            
            # 只处理存在的列
            if 'size' in df.columns:
                df['size'] = df['size'].fillna('').astype(str).str.strip()
            else:
                df['size'] = ''  # 添加默认的size列
                
            if 'color' in df.columns:
                df['color'] = df['color'].fillna('').astype(str).str.strip()
            else:
                df['color'] = ''  # 添加默认的color列
                
            df['standard_unit_price'] = pd.to_numeric(df['standard_unit_price'], errors='coerce')
            
            # 创建复合键用于快速查找
            # 根据可用列创建复合键
            df['composite_key'] = df['item_id'].astype(str)  # 始终使用item_id
            
            # 添加size到复合键
            df['composite_key'] = df['composite_key'] + '|' + df['size'].astype(str)
            
            # 添加color到复合键
            df['composite_key'] = df['composite_key'] + '|' + df['color'].astype(str)
            
            return df
            
        except Exception as e:
            logger.error(f"加载规格数据失败: {str(e)}")
            return None
            
    def compare_orders(self, order_file_path, spec_file_path, check_total_calc=True):
        """
        比对订单与产品规格
        
        Args:
            order_file_path: 订单Excel文件路径
            spec_file_path: 产品规格Excel文件路径
            check_total_calc: 是否检查总价计算
            
        Returns:
            dict: 比对结果，包含结果文件路径和统计信息
        """
        try:
            # 加载数据
            order_df = self.load_order_data(order_file_path)
            spec_df = self.load_spec_data(spec_file_path)
            
            if order_df is None or spec_df is None:
                return {'error': '数据加载失败'}
                
            # 创建规格数据的查找字典
            spec_dict = {}
            for _, row in spec_df.iterrows():
                # 安全地获取列值
                item_id = str(row.get('item_id', '')).strip()
                size = str(row.get('size', '')).strip()
                color = str(row.get('color', '')).strip()
                
                key = f"{item_id}|{size}|{color}"
                spec_dict[key] = {
                    'standard_unit_price': row.get('standard_unit_price', 0),
                    'product_name': row.get('product_name', ''),
                    'description': row.get('description', '')
                }
            
            # 初始化结果列
            order_df['核对状态'] = '通过'
            order_df['错误详情'] = ''
            
            # 记录统计信息
            stats = {
                'total_records': len(order_df),
                'error_records': 0,
                'error_types': {error_type: 0 for error_type in self.ERROR_TYPES.keys()}
            }
            
            # 逐行比对
            for idx, row in order_df.iterrows():
                errors = []
                
                # 检查必需字段是否为空
                if pd.isna(row['item_id']) or str(row['item_id']).strip() == '':
                    errors.append('产品ID为空')
                # 检查价格字段（可能在不同的列中）
                elif self._get_price_value(row) <= 0:
                    errors.append('单价无效')
                else:
                    # 构造查找键
                    item_id = str(row.get('item_id', '')).strip()
                    size = str(row.get('size', '')).strip()
                    color = str(row.get('color', '')).strip()
                    
                    lookup_key = f"{item_id}|{size}|{color}"
                    
                    if lookup_key not in spec_dict:
                        # 检查是否只是产品ID不存在
                        item_exists = any(str(spec_row.get('item_id', '')).strip() == str(row.get('item_id', '')).strip() for _, spec_row in spec_df.iterrows())
                        
                        if not item_exists:
                            errors.append(self.ERROR_TYPES['PRODUCT_NOT_FOUND'])
                            stats['error_types']['PRODUCT_NOT_FOUND'] += 1
                        else:
                            # 产品ID存在，但尺寸或颜色不匹配
                            item_specs = spec_df[spec_df['item_id'] == row.get('item_id', '')]
                            
                            # 检查尺寸是否匹配
                            order_size = str(row.get('size', '')).strip()
                            if order_size:  # 只有当订单中有尺寸时才检查
                                size_match = any(str(spec_row.get('size', '')).strip() == order_size for _, spec_row in item_specs.iterrows())
                            else:
                                size_match = True  # 如果订单中没有尺寸，则视为匹配
                            
                            # 检查颜色是否匹配
                            order_color = str(row.get('color', '')).strip()
                            if order_color:  # 只有当订单中有颜色时才检查
                                color_match = any(str(spec_row.get('color', '')).strip() == order_color for _, spec_row in item_specs.iterrows())
                            else:
                                color_match = True  # 如果订单中没有颜色，则视为匹配
                            
                            if not size_match:
                                errors.append(self.ERROR_TYPES['SIZE_MISMATCH'])
                                stats['error_types']['SIZE_MISMATCH'] += 1
                                
                            if not color_match:
                                errors.append(self.ERROR_TYPES['COLOR_MISMATCH'])
                                stats['error_types']['COLOR_MISMATCH'] += 1
                    else:
                        # 找到匹配的规格，检查单价
                        spec_info = spec_dict[lookup_key]
                        
                        # 检查单价是否匹配
                        if 'standard_unit_price' in spec_info:
                            try:
                                order_price = self._get_price_value(row)
                                spec_price = float(spec_info['standard_unit_price'])
                                if order_price > 0 and abs(order_price - spec_price) > 0.01:  # 允许0.01的误差
                                    errors.append(f"{self.ERROR_TYPES['PRICE_MISMATCH']} (标准价格: {spec_price})")
                                    stats['error_types']['PRICE_MISMATCH'] += 1
                            except (ValueError, TypeError):
                                errors.append(f"单价格式无效")
                                stats['error_types']['PRICE_MISMATCH'] += 1
                            
                        # 检查总价计算（如果有相关列）
                        if check_total_calc and self._has_quantity_and_total(row):
                            try:
                                unit_price = self._get_price_value(row)
                                quantity = self._get_quantity_value(row)
                                total_price = self._get_total_value(row)
                                
                                if unit_price > 0 and quantity > 0 and total_price > 0:
                                    expected_total = unit_price * quantity
                                    if abs(total_price - expected_total) > 0.01:
                                        errors.append(f"{self.ERROR_TYPES['TOTAL_CALC_ERROR']} (应为: {expected_total:.2f})")
                                        stats['error_types']['TOTAL_CALC_ERROR'] += 1
                            except (ValueError, TypeError):
                                # 如果转换失败，跳过总价检查
                                pass
                
                # 更新结果
                if errors:
                    order_df.at[idx, '核对状态'] = '有问题'
                    order_df.at[idx, '错误详情'] = '; '.join(errors)
                    stats['error_records'] += 1
                    
            # 生成结果文件
            result_file_id = str(uuid.uuid4())
            result_filename = f"order_comparison_{result_file_id}.xlsx"
            result_file_path = os.path.join(self.output_dir, result_filename)
            
            # 保存到Excel并添加格式
            self.save_with_formatting(order_df, result_file_path)
            
            return {
                'result_file_id': result_file_id,
                'result_file_path': result_file_path,
                'filename': result_filename,
                'stats': stats
            }
            
        except Exception as e:
            logger.error(f"订单比对失败: {str(e)}")
            return {'error': f'比对失败: {str(e)}'}
            
    def save_with_formatting(self, df, file_path):
        """
        保存DataFrame到Excel并添加格式化，支持多工作表
        
        Args:
            df: 要保存的DataFrame
            file_path: 输出文件路径
        """
        try:
            # 检查是否有多个工作表
            if '工作表' in df.columns and '表格序号' in df.columns:
                # 多工作表模式
                self._save_multi_sheet_with_formatting(df, file_path)
            else:
                # 单工作表模式
                self._save_single_sheet_with_formatting(df, file_path)
                
        except Exception as e:
            logger.error(f"Excel格式化失败: {str(e)}")
            # 如果格式化失败，至少保证基本的Excel文件可用
            df.to_excel(file_path, index=False)
    
    def _save_single_sheet_with_formatting(self, df, file_path):
        """保存单工作表并格式化"""
        # 先保存为Excel
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        # 重新打开进行格式化
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 找到状态列和错误详情列的索引
        status_col_idx = None
        error_col_idx = None
        
        for col_idx, cell in enumerate(ws[1], 1):  # 第一行是标题行
            if cell.value == '核对状态':
                status_col_idx = col_idx
            elif cell.value == '错误详情':
                error_col_idx = col_idx
                
        # 格式化有问题的行
        if status_col_idx:
                for row_idx in range(2, ws.max_row + 1):  # 从第二行开始（跳过标题）
                    status_cell = ws.cell(row=row_idx, column=status_col_idx)
                    
                    if status_cell.value == '有问题':
                        # 高亮整行
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.fill = self.ERROR_FILL
                            
                        # 状态列使用红色字体
                        status_cell.font = self.ERROR_FONT
                        
                            # 如果有错误详情列，添加批注
                    if error_col_idx:
                        error_cell = ws.cell(row=row_idx, column=error_col_idx)
                        if error_cell.value:
                            comment = Comment(str(error_cell.value), "系统")
                            error_cell.comment = comment
                            
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # 保存格式化后的文件
        wb.save(file_path)
    
    def _save_multi_sheet_with_formatting(self, df, file_path):
        """保存多工作表并格式化"""
        # 按工作表分组
        sheet_groups = df.groupby('工作表')
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, sheet_df in sheet_groups:
                # 移除工作表标识列
                sheet_df_clean = sheet_df.drop(['工作表', '表格序号'], axis=1)
                
                # 保存到对应的工作表
                sheet_df_clean.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 格式化当前工作表
                self._format_worksheet(writer.sheets[sheet_name], sheet_df_clean)
        
        logger.info(f"保存了 {len(sheet_groups)} 个工作表到 {file_path}")
    
    def _format_worksheet(self, ws, df):
        """格式化单个工作表"""
        # 找到状态列和错误详情列的索引
        status_col_idx = None
        error_col_idx = None
        
        for col_idx, cell in enumerate(ws[1], 1):  # 第一行是标题行
            if cell.value == '核对状态':
                status_col_idx = col_idx
            elif cell.value == '错误详情':
                error_col_idx = col_idx
                
        # 格式化有问题的行
        if status_col_idx:
            for row_idx in range(2, len(df) + 2):  # 从第二行开始（跳过标题行）
                status_cell = ws.cell(row=row_idx, column=status_col_idx)
                if status_cell.value == '有问题':
                    # 高亮整行
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = self.ERROR_FILL
                        if col_idx == status_col_idx or col_idx == error_col_idx:
                            cell.font = self.ERROR_FONT
                            
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def get_comparison_summary(self, stats):
        """
        生成比对摘要报告
        
        Args:
            stats: 统计信息字典
            
        Returns:
            str: 格式化的摘要报告
        """
        total = stats['total_records']
        errors = stats['error_records']
        success_rate = ((total - errors) / total * 100) if total > 0 else 100
        
        summary = f"""
比对摘要报告
=============
总记录数: {total}
通过记录数: {total - errors}
问题记录数: {errors}
通过率: {success_rate:.1f}%

错误类型统计:
"""
        
        for error_type, count in stats['error_types'].items():
            if count > 0:
                error_name = self.ERROR_TYPES.get(error_type, error_type)
                summary += f"- {error_name}: {count}条\n"
                
        return summary

